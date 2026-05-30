import os
import re
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

cards_dir = r"scripts/cards"
cards_json_path = r"Arknights_Mizuki/localization/zhs/cards.json"

# Load localization
with open(cards_json_path, "r", encoding="utf-8") as f:
    loc_data = json.load(f)

# CardType mapping
card_type_map = {
    "CardType.Attack": "攻击",
    "CardType.Skill": "技能",
    "CardType.Power": "能力",
    "(CardType)3": "能力",
}

# CardRarity mapping
card_rarity_map = {
    "CardRarity.Basic": "基础",
    "CardRarity.Common": "普通",
    "CardRarity.Uncommon": "罕见",
    "CardRarity.Rare": "稀有",
    "(CardRarity)4": "远古",
    "CardRarity.Ancient": "远古",
}

# Keyword name -> display name
keyword_display = {
    "AutoPlay.Autoplay": "自动打出",
    "CardKeyword.Exhaust": "消耗",
    "CardKeyword.Ethereal": "虚无",
    "CardKeyword.Innate": "固有",
    "(CardKeyword)1": "消耗",
    "(CardKeyword)2": "虚无",
    "(CardKeyword)5": "自动打出",
}

# DynamicVar type -> display name
dv_name_map = {
    "Damage": "伤害",
    "Block": "格挡",
    "SanityPower": "损伤",
    "VulnerablePower": "易伤",
    "StealthPower": "潜行",
    "SanityBuffPower": "攻击附带损伤",
    "SanityThornsPower": "毒素棘刺",
    "AttackApplySanityPower": "创伤性癔症",
    "SanityUnlimitPower": "损伤解限",
    "ErodeTidePower": "侵蚀",
    "WeakPower": "虚弱",
    "VigorPower": "力量(下一击)",
    "SheildPower": "格挡(急速)",
    "Cards": "抽牌",
    "Energy": "能量",
    "Heal": "回复",
    "Repeat": "次数",
    "Times": "次数",
    "DiscardPicks": "选择数量",
    "HpLoss": "失去生命",
}


def parse_cs_file(filepath):
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()

    # Class name
    class_match = re.search(r'public class (\w+)\s*:\s*CustomCardModel', content)
    class_name = class_match.group(1) if class_match else ""

    # Card type
    type_match = re.search(r'CardType\s+type\s*=\s*(CardType\.\w+|\(CardType\)\d+)', content)
    card_type_raw = type_match.group(1) if type_match else ""
    card_type = card_type_map.get(card_type_raw, card_type_raw)

    # Card rarity
    rarity_match = re.search(r'CardRarity\s+rarity\s*=\s*(CardRarity\.\w+|\(CardRarity\)\d+)', content)
    rarity_raw = rarity_match.group(1) if rarity_match else ""
    rarity = card_rarity_map.get(rarity_raw, rarity_raw)

    # Energy cost
    energy_match = re.search(r'energyCost\s*=\s*(\d+)', content)
    energy = int(energy_match.group(1)) if energy_match else 0

    # Keywords
    keywords = []
    kw_block = re.search(r'CanonicalKeywords\s*=>(.*?);', content, re.DOTALL)
    if kw_block:
        kw_text = kw_block.group(1)
        # Match AutoPlay.Autoplay
        for m in re.finditer(r'AutoPlay\.Autoplay', kw_text):
            keywords.append("自动打出")
        # Match CardKeyword.Xxx
        for m in re.finditer(r'CardKeyword\.(\w+)', kw_text):
            kw_name = m.group(1)
            if kw_name != "Exhaust":
                keywords.append(kw_name)
        # Match (CardKeyword)N
        for m in re.finditer(r'\(CardKeyword\)(\d+)', kw_text):
            k = f"(CardKeyword){m.group(1)}"
            if k in keyword_display:
                keywords.append(keyword_display[k])
        # Exhaust
        if re.search(r'CardKeyword\.Exhaust', kw_text):
            keywords.append("消耗")

    # CanonicalVars - parse base values
    base_values = {}  # var_name -> base_value
    cv_block = re.search(r'CanonicalVars\s*=>(.*?);', content, re.DOTALL)
    if cv_block:
        cv_text = cv_block.group(1)
        # Parse DamageVar(baseValue, ...)
        for m in re.finditer(r'DamageVar\((\d+)m', cv_text):
            base_values["Damage"] = int(m.group(1))
        # Parse BlockVar(baseValue, ...)
        for m in re.finditer(r'BlockVar\((\d+)m', cv_text):
            base_values["Block"] = int(m.group(1))
        # Parse PowerVar<Type>(baseValue)
        for m in re.finditer(r'PowerVar<(\w+)>\((\d+)m\)', cv_text):
            power_name = m.group(1)
            base_values[power_name] = int(m.group(2))
        # Parse CardsVar(baseValue)
        for m in re.finditer(r'CardsVar\((\d+)\)', cv_text):
            base_values["Cards"] = int(m.group(1))
        # Parse EnergyVar(baseValue)
        for m in re.finditer(r'EnergyVar\((\d+)\)', cv_text):
            base_values["Energy"] = int(m.group(1))
        # Parse HealVar(baseValue)
        for m in re.finditer(r'HealVar\((\d+)m\)', cv_text):
            base_values["Heal"] = int(m.group(1))
        # Parse RepeatVar(baseValue)
        for m in re.finditer(r'RepeatVar\((\d+)\)', cv_text):
            base_values["Repeat"] = int(m.group(1))
        # Parse IntVar("name", baseValue)
        for m in re.finditer(r'IntVar\("(\w+)",\s*(\d+)m\)', cv_text):
            base_values[m.group(1)] = int(m.group(2))

    # OnUpgrade
    upgrade_list = []
    up_block = re.search(r'protected override void OnUpgrade\(\)\s*\{(.*?)\n\s*\}', content, re.DOTALL)
    if up_block:
        up_text = up_block.group(1)
        # UpgradeValueBy
        for m in re.finditer(r'DynamicVars\["(\w+)"\]\.UpgradeValueBy\((-?\d+)m?\)', up_text):
            var_name = m.group(1)
            delta = int(m.group(2))
            disp = dv_name_map.get(var_name, var_name)
            if delta > 0:
                upgrade_list.append(f"{disp}+{delta}")
            else:
                upgrade_list.append(f"{disp}{delta}")
        for m in re.finditer(r'DynamicVars\.(\w+)\.UpgradeValueBy\((-?\d+)m?\)', up_text):
            var_name = m.group(1)
            delta = int(m.group(2))
            disp = dv_name_map.get(var_name, var_name)
            if delta > 0:
                upgrade_list.append(f"{disp}+{delta}")
            else:
                upgrade_list.append(f"{disp}{delta}")
        for m in re.finditer(r'DynamicVars\.(\w+)\.UpgradeValueBy\((-?\d+)\)', up_text):
            var_name = m.group(1)
            delta = int(m.group(2))
            disp = dv_name_map.get(var_name, var_name)
            if delta > 0:
                upgrade_list.append(f"{disp}+{delta}")
            else:
                upgrade_list.append(f"{disp}{delta}")

        # AddKeyword
        for m in re.finditer(r'AddKeyword\(CardKeyword\.(\w+)\)', up_text):
            kw = m.group(1)
            upgrade_list.append(f"获得{kw}")
        # RemoveKeyword
        for m in re.finditer(r'RemoveKeyword\(CardKeyword\.(\w+)\)', up_text):
            kw = m.group(1)
            upgrade_list.append(f"移除{kw}")
        # EnergyCost.UpgradeBy
        for m in re.finditer(r'EnergyCost\.UpgradeBy\((-?\d+)\)', up_text):
            delta = int(m.group(1))
            if delta < 0:
                upgrade_list.append(f"耗能{delta}")

    return class_name, card_type, rarity, energy, keywords, base_values, upgrade_list


def find_localization_key(class_name):
    name_to_key = {
        "MzkStrike": "ARKNIGHTS_MIZUKI-MZK_STRIKE",
        "MzkDefence": "ARKNIGHTS_MIZUKI-MZK_DEFENCE",
        "MzkNeurotoxin": "ARKNIGHTS_MIZUKI-MZK_NEUROTOXIN",
        "Awaken": "ARKNIGHTS_MIZUKI-AWAKEN",
        "MzkUnlimit": "ARKNIGHTS_MIZUKI-MZK_UNLIMIT",
        "SanityBurst": "ARKNIGHTS_MIZUKI-MZK_SANITY_BURST",
        "SanityThorns": "ARKNIGHTS_MIZUKI-MZK_SANITY_THORNS",
        "SeaSnakeBite": "ARKNIGHTS_MIZUKI-SEA_SNAKE_BITE",
        "SeaSnakeBigBite": "ARKNIGHTS_MIZUKI-SEA_SNAKE_BIG_BITE",
        "MzkStealth": "ARKNIGHTS_MIZUKI-MZK_STEALTH",
        "MzkEnergyDrain": "ARKNIGHTS_MIZUKI-MZK_ENERGY_DRAIN",
        "MzkMirrorFlower": "ARKNIGHTS_MIZUKI-MZK_MIRROR_FLOWER",
        "Attack_Sanity": "ARKNIGHTS_MIZUKI-ATTACK_SANITY",
        "ErodeTide": "ARKNIGHTS_MIZUKI-ERODE_TIDE",
        "BlueSeed": "ARKNIGHTS_MIZUKI-BLUE_SEED",
        "Learn": "ARKNIGHTS_MIZUKI-LEARN",
        "Share": "ARKNIGHTS_MIZUKI-SHARE",
        "Shock": "ARKNIGHTS_MIZUKI-SHOCK",
        "Explain": "ARKNIGHTS_MIZUKI-EXPLAIN",
        "SpeedUp": "ARKNIGHTS_MIZUKI-SPEED_UP",
        "IronWave": "ARKNIGHTS_MIZUKI-IRON_WAVE",
        "Trauma": "ARKNIGHTS_MIZUKI-TRAUMA",
        "Sign": "ARKNIGHTS_MIZUKI-SIGN",
        "Return": "ARKNIGHTS_MIZUKI-RETURN",
        "Whirlpool": "ARKNIGHTS_MIZUKI-WHIRLPOOL",
        "Escape": "ARKNIGHTS_MIZUKI-ESCAPE",
        "Revive": "ARKNIGHTS_MIZUKI-REVIVE",
        "AwakenPlus": "ARKNIGHTS_MIZUKI-AWAKEN_PLUS",
        "Hurt": "ARKNIGHTS_MIZUKI-HURT",
        "Overload": "ARKNIGHTS_MIZUKI-OVERLOAD",
        "Shadow": "ARKNIGHTS_MIZUKI-SHADOW",
        "Sweep": "ARKNIGHTS_MIZUKI-SWEEP",
        "Bigge": "ARKNIGHTS_MIZUKI-BIGGE",
        "CutDown": "ARKNIGHTS_MIZUKI-CUT_DOWN",
        "Food": "ARKNIGHTS_MIZUKI-FOOD",
        "ValStrike": "ARKNIGHTS_MIZUKI-VAL_STRIKE",
        "Around": "ARKNIGHTS_MIZUKI-AROUND",
        "Spray": "ARKNIGHTS_MIZUKI-SPRAY",
        "AutoSheild": "ARKNIGHTS_MIZUKI-AUTO_SHEILD",
        "AutoPlayAll": "ARKNIGHTS_MIZUKI-AUTO_PLAY_ALL",
        "EchoAttack": "ARKNIGHTS_MIZUKI-ECHO_ATTACK",
        "EchoSanity": "ARKNIGHTS_MIZUKI-ECHO_SANITY",
        "EchoGenerateAuto": "ARKNIGHTS_MIZUKI-ECHO_GENERATE_AUTO",
        "End": "ARKNIGHTS_MIZUKI-END",
        "Copy": "ARKNIGHTS_MIZUKI-COPY",
        "SuperSpeed": "ARKNIGHTS_MIZUKI-SUPER_SPEED",
        "AdaptPain": "ARKNIGHTS_MIZUKI-ADAPT_PAIN",
    }
    return name_to_key.get(class_name, "")


def get_name_and_desc(key):
    name = loc_data.get(f"{key}.title", "")
    desc = loc_data.get(f"{key}.description", "")
    return name, desc


def replace_vars_in_desc(desc, base_values):
    """Replace {Var:diff()} and {Var:energyIcons()} with actual numbers."""
    # Replace {Var:diff()} with actual base value
    for var_name, val in base_values.items():
        desc = desc.replace(f"{{{var_name}:diff()}}", str(val))
        desc = desc.replace(f"{{{var_name}:inverseDiff()}}", str(val))
        desc = desc.replace(f"{{{var_name}:energyIcons()}}", f"[E]")
    return desc


def clean_markup(desc):
    """Clean color markup tags."""
    desc = desc.replace("[gold]", "").replace("[/gold]", "")
    desc = desc.replace("[blue]", "").replace("[/blue]", "")
    desc = desc.replace("[red]", "").replace("[/red]", "")
    desc = desc.replace("[green]", "").replace("[/green]", "")
    return desc


def main():
    files = sorted([f for f in os.listdir(cards_dir) if f.endswith('.cs')])

    results = []
    for fname in files:
        fpath = os.path.join(cards_dir, fname)
        class_name, card_type, rarity, energy, keywords, base_values, upgrade_list = parse_cs_file(fpath)
        if not class_name:
            continue

        key = find_localization_key(class_name)
        name, desc = get_name_and_desc(key) if key else ("", "")
        if not name:
            name = class_name  # fallback

        # Replace variables in description with actual numbers
        desc_filled = replace_vars_in_desc(desc, base_values) if desc else ""
        desc_clean = clean_markup(desc_filled)

        # Build keywords string
        kw_str = "、".join(keywords) if keywords else ""

        # Build upgrade string
        up_str = "，".join(upgrade_list) if upgrade_list else ""

        # Build base values string
        bv_parts = []
        for var_name, val in base_values.items():
            disp = dv_name_map.get(var_name, var_name)
            bv_parts.append(f"{disp}{val}")
        bv_str = "，".join(bv_parts) if bv_parts else ""

        results.append({
            "file": fname,
            "class": class_name,
            "name": name,
            "type": card_type,
            "rarity": rarity,
            "energy": energy,
            "keywords": kw_str,
            "base_values": bv_str,
            "upgrade": up_str,
            "desc": desc_clean,
        })

    # Sort by rarity then name
    rarity_order = {"基础": 0, "普通": 1, "罕见": 2, "稀有": 3, "远古": 4}
    results.sort(key=lambda r: (rarity_order.get(r["rarity"], 99), r["name"]))

    # Build markdown
    lines = []
    lines.append("| # | 名称 | 类别 | 稀有度 | 耗能 | 关键词 | 基础数值 | 升级效果 | 描述 |")
    lines.append("|---|------|------|--------|------|--------|----------|----------|------|")
    for i, r in enumerate(results, 1):
        lines.append(
            f"| {i} | {r['name']} | {r['type']} | {r['rarity']} | {r['energy']} | {r['keywords']} | {r['base_values']} | {r['upgrade']} | {r['desc']} |"
        )
    lines.append(f"\n共 {len(results)} 张卡牌")

    output = "\n".join(lines)
    print(output)

    with open("card_list.md", "w", encoding="utf-8") as f:
        f.write(output)
    print("\n已写入 card_list.md")

    # Generate xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = "卡牌列表"

    # Styles
    header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Microsoft YaHei", size=10)
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    rarity_fills = {
        "基础": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "普通": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        "罕见": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
        "稀有": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "远古": PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"),
    }

    headers = ["#", "名称", "类别", "稀有度", "耗能", "关键词", "基础数值", "升级效果", "描述"]
    col_widths = [5, 14, 8, 8, 6, 16, 24, 22, 65]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for i, r in enumerate(results, 1):
        row = i + 1
        row_data = [i, r["name"], r["type"], r["rarity"], r["energy"], r["keywords"], r["base_values"], r["upgrade"], r["desc"]]
        row_fill = rarity_fills.get(r["rarity"])

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_alignment if col_idx != 1 else Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill

    for row in range(2, len(results) + 2):
        ws.row_dimensions[row].height = 36

    ws.row_dimensions[1].height = 24

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(results) + 1}"

    wb.save("card_list.xlsx")
    print("已写入 card_list.xlsx")


if __name__ == "__main__":
    main()